"""
Export Service - PDF and Excel report generation
"""
from datetime import datetime
from typing import List, Dict, Any
from io import BytesIO


def generate_pdf_report(traffic_data: List[Dict[str, Any]]) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        elements.append(Paragraph("Traffic Analysis Report", styles["Title"]))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["Normal"]))
        elements.append(Spacer(1, 12))
        
        elements.append(Paragraph("Traffic Data Summary", styles["Heading2"]))
        
        table_data = [["Location", "Flow Rate", "Vehicles", "Speed", "Congestion", "Timestamp"]]
        for data in traffic_data:
            table_data.append([
                data.get("location_name", "N/A"),
                f"{data.get('flow_rate', 0):.1f}",
                str(data.get("vehicle_count", 0)),
                f"{data.get('avg_speed', 0):.1f} km/h",
                data.get("congestion_level", "N/A"),
                data.get("timestamp", "N/A"),
            ])
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)
        
        doc.build(elements)
        return buffer.getvalue()
    except ImportError:
        return _generate_simple_report(traffic_data)


def generate_excel_report(traffic_data: List[Dict[str, Any]]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Traffic Data"
        
        headers = ["Location", "Latitude", "Longitude", "Flow Rate", "Vehicles", "Speed (km/h)", "Congestion", "Timestamp"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        for row, data in enumerate(traffic_data, 2):
            ws.cell(row=row, column=1, value=data.get("location_name", ""))
            ws.cell(row=row, column=2, value=data.get("latitude", 0))
            ws.cell(row=row, column=3, value=data.get("longitude", 0))
            ws.cell(row=row, column=4, value=data.get("flow_rate", 0))
            ws.cell(row=row, column=5, value=data.get("vehicle_count", 0))
            ws.cell(row=row, column=6, value=data.get("avg_speed", 0))
            ws.cell(row=row, column=7, value=data.get("congestion_level", ""))
            ws.cell(row=row, column=8, value=str(data.get("timestamp", "")))
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].auto_size = True
        
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
    except ImportError:
        return _generate_csv_report(traffic_data)


def _generate_simple_report(traffic_data: List[Dict[str, Any]]) -> bytes:
    content = "Traffic Analysis Report\n"
    content += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
    content += "Location,Flow Rate,Vehicles,Speed,Congestion\n"
    for data in traffic_data:
        content += f"{data.get('location_name', '')},{data.get('flow_rate', 0)},{data.get('vehicle_count', 0)},{data.get('avg_speed', 0)},{data.get('congestion_level', '')}\n"
    return content.encode("utf-8")


def _generate_csv_report(traffic_data: List[Dict[str, Any]]) -> bytes:
    return _generate_simple_report(traffic_data)


def generate_site_pdf_report(report_data: Dict[str, Any]) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        elements.append(Paragraph("Site Analysis Report", styles["Title"]))
        elements.append(Paragraph(f"Location: {report_data.get('location', 'N/A')}", styles["Normal"]))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["Normal"]))
        elements.append(Spacer(1, 12))
        
        terrain = report_data.get("terrain", {})
        if terrain:
            elements.append(Paragraph("Terrain & Environmental Data", styles["Heading2"]))
            terrain_data = [
                ["Elevation (m)", "Soil Type", "Flood Risk", "Seismic Zone", "Slope (°)"],
                [
                    str(terrain.get("elevation_m", "N/A")),
                    str(terrain.get("soil_type", "N/A")),
                    str(terrain.get("flood_risk", "N/A")),
                    str(terrain.get("seismic_zone", "N/A")),
                    str(terrain.get("slope_degrees", "N/A")),
                ]
            ]
            table = Table(terrain_data)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 12))
        
        population = report_data.get("population", {})
        if population:
            elements.append(Paragraph("Demographic Data", styles["Heading2"]))
            pop_data = [
                ["Population", "Density/sqkm", "Growth Rate %", "Households"],
                [
                    str(population.get("current_population", "N/A")),
                    str(population.get("density_per_sqkm", "N/A")),
                    str(population.get("growth_rate", "N/A")),
                    str(population.get("households", "N/A")),
                ]
            ]
            table = Table(pop_data)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 12))
        
        land_use = report_data.get("land_use", {})
        if land_use:
            elements.append(Paragraph("Land Use & Zoning", styles["Heading2"]))
            land_data = [
                ["Urban %", "Commercial %", "Industrial %", "Residential %", "Avg Price/sqft"],
                [
                    str(land_use.get("urban_area_percent", "N/A")),
                    str(land_use.get("commercial_percent", "N/A")),
                    str(land_use.get("industrial_percent", "N/A")),
                    str(land_use.get("residential_percent", "N/A")),
                    f"₹{land_use.get('avg_land_price_per_sqft', 'N/A')}",
                ]
            ]
            table = Table(land_data)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.darkgreen),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 12))
        
        risks = report_data.get("risks", {})
        if risks:
            elements.append(Paragraph("Risk Assessment", styles["Heading2"]))
            risk_list = risks.get("risks", [])
            if risk_list:
                risk_headers = ["Risk Type", "Probability", "Severity", "Mitigation"]
                risk_rows = [risk_headers]
                for risk in risk_list:
                    risk_rows.append([
                        str(risk.get("type", "N/A")),
                        f"{risk.get('probability', 0):.2f}",
                        str(risk.get("severity", "N/A")),
                        str(risk.get("mitigation", "N/A"))[:30] + "..." if len(str(risk.get("mitigation", ""))) > 30 else str(risk.get("mitigation", "N/A")),
                    ])
                table = Table(risk_rows)
                table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.darkred),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 8),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]))
                elements.append(table)
            else:
                elements.append(Paragraph("No significant risks identified.", styles["Normal"]))
            elements.append(Spacer(1, 12))
        
        recommendation = report_data.get("recommendation", "")
        if recommendation:
            elements.append(Paragraph("AI Recommendations", styles["Heading2"]))
            for line in recommendation.split("\n"):
                if line.strip():
                    elements.append(Paragraph(line, styles["Normal"]))
        
        doc.build(elements)
        return buffer.getvalue()
    except ImportError:
        return _generate_simple_site_report(report_data)


def generate_site_excel_report(report_data: Dict[str, Any]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        
        ws = wb.active
        ws.title = "Site Overview"
        ws.cell(row=1, column=1, value="Site Analysis Report").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Location: {report_data.get('location', 'N/A')}")
        ws.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=4, column=1, value=f"Overall Score: {report_data.get('overall_score', 'N/A')}/10")
        
        terrain = report_data.get("terrain", {})
        if terrain:
            ws_terrain = wb.create_sheet("Terrain & Environment")
            headers = ["Attribute", "Value", "Unit"]
            for col, header in enumerate(headers, 1):
                cell = ws_terrain.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            terrain_items = [
                ("Elevation", terrain.get("elevation_m"), "m"),
                ("Soil Type", terrain.get("soil_type"), ""),
                ("Flood Risk", terrain.get("flood_risk"), ""),
                ("Seismic Zone", terrain.get("seismic_zone"), ""),
                ("Slope", terrain.get("slope_degrees"), "degrees"),
                ("Ground Water Depth", terrain.get("ground_water_depth_m"), "m"),
                ("Air Quality Index", terrain.get("air_quality_index"), ""),
            ]
            for row, (attr, val, unit) in enumerate(terrain_items, 2):
                ws_terrain.cell(row=row, column=1, value=attr)
                ws_terrain.cell(row=row, column=2, value=str(val) if val else "N/A")
                ws_terrain.cell(row=row, column=3, value=unit)
        
        population = report_data.get("population", {})
        if population:
            ws_pop = wb.create_sheet("Demographics")
            headers = ["Metric", "Value"]
            for col, header in enumerate(headers, 1):
                cell = ws_pop.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            pop_items = [
                ("Current Population", population.get("current_population")),
                ("Density per sqkm", population.get("density_per_sqkm")),
                ("Growth Rate %", population.get("growth_rate")),
                ("Households", population.get("households")),
                ("Avg Household Size", population.get("avg_household_size")),
            ]
            for row, (metric, val) in enumerate(pop_items, 2):
                ws_pop.cell(row=row, column=1, value=metric)
                ws_pop.cell(row=row, column=2, value=str(val) if val else "N/A")
        
        land_use = report_data.get("land_use", {})
        if land_use:
            ws_land = wb.create_sheet("Land Use")
            headers = ["Category", "Percentage", "Value"]
            for col, header in enumerate(headers, 1):
                cell = ws_land.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            land_items = [
                ("Urban Area", land_use.get("urban_area_percent"), "%"),
                ("Commercial", land_use.get("commercial_percent"), "%"),
                ("Industrial", land_use.get("industrial_percent"), "%"),
                ("Residential", land_use.get("residential_percent"), "%"),
                ("Agricultural", land_use.get("agricultural_percent"), "%"),
                ("Open Spaces", land_use.get("open_spaces_percent"), "%"),
                ("Avg Land Price", land_use.get("avg_land_price_per_sqft"), "₹/sqft"),
            ]
            for row, (cat, pct, unit) in enumerate(land_items, 2):
                ws_land.cell(row=row, column=1, value=cat)
                ws_land.cell(row=row, column=2, value=str(pct) if pct else "N/A")
                ws_land.cell(row=row, column=3, value=unit)
        
        risks = report_data.get("risks", {})
        if risks and risks.get("risks"):
            ws_risk = wb.create_sheet("Risk Assessment")
            headers = ["Risk Type", "Probability", "Severity", "Mitigation"]
            for col, header in enumerate(headers, 1):
                cell = ws_risk.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            for row, risk in enumerate(risks.get("risks", []), 2):
                ws_risk.cell(row=row, column=1, value=risk.get("type", ""))
                ws_risk.cell(row=row, column=2, value=risk.get("probability", 0))
                ws_risk.cell(row=row, column=3, value=risk.get("severity", ""))
                ws_risk.cell(row=row, column=4, value=risk.get("mitigation", ""))
        
        recommendation = report_data.get("recommendation", "")
        if recommendation:
            ws_rec = wb.create_sheet("Recommendations")
            ws_rec.cell(row=1, column=1, value="AI Recommendations").font = Font(bold=True)
            for row, line in enumerate(recommendation.split("\n"), 2):
                ws_rec.cell(row=row, column=1, value=line)
        
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
    except ImportError:
        return _generate_csv_site_report(report_data)


def _generate_simple_site_report(report_data: Dict[str, Any]) -> bytes:
    content = "Site Analysis Report\n"
    content += f"Location: {report_data.get('location', 'N/A')}\n"
    content += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
    content += f"Overall Score: {report_data.get('overall_score', 'N/A')}/10\n\n"
    
    terrain = report_data.get("terrain", {})
    if terrain:
        content += "=== TERRAIN & ENVIRONMENT ===\n"
        content += f"Elevation: {terrain.get('elevation_m', 'N/A')}m\n"
        content += f"Soil Type: {terrain.get('soil_type', 'N/A')}\n"
        content += f"Flood Risk: {terrain.get('flood_risk', 'N/A')}\n\n"
    
    population = report_data.get("population", {})
    if population:
        content += "=== DEMOGRAPHICS ===\n"
        content += f"Population: {population.get('current_population', 'N/A')}\n"
        content += f"Density: {population.get('density_per_sqkm', 'N/A')}/sqkm\n\n"
    
    land_use = report_data.get("land_use", {})
    if land_use:
        content += "=== LAND USE ===\n"
        content += f"Urban: {land_use.get('urban_area_percent', 'N/A')}%\n"
        content += f"Commercial: {land_use.get('commercial_percent', 'N/A')}%\n\n"
    
    if report_data.get("recommendation"):
        content += "=== RECOMMENDATIONS ===\n"
        content += report_data.get("recommendation", "")
    
    return content.encode("utf-8")


def _generate_csv_site_report(report_data: Dict[str, Any]) -> bytes:
    return _generate_simple_site_report(report_data)


def generate_forecast_pdf_report(forecast_data: List[Dict[str, Any]]) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        elements.append(Paragraph("Traffic Forecast Report", styles["Title"]))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["Normal"]))
        elements.append(Spacer(1, 12))
        
        elements.append(Paragraph("Forecast Data Summary", styles["Heading2"]))
        
        table_data = [["Location", "Type", "Base Volume", "Predicted Volume", "Growth Rate", "Confidence", "Status"]]
        for data in forecast_data:
            table_data.append([
                data.get("location_name", "N/A"),
                data.get("forecast_type", "N/A"),
                str(data.get("base_traffic_volume", 0)),
                str(data.get("predicted_traffic_volume", "N/A")),
                f"{data.get('growth_rate', 0):.2f}%" if data.get("growth_rate") else "N/A",
                f"{data.get('confidence_score', 0):.2f}" if data.get("confidence_score") else "N/A",
                data.get("status", "N/A"),
            ])
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)
        
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("AI Insights", styles["Heading2"]))
        for data in forecast_data:
            if data.get("ai_insights"):
                elements.append(Paragraph(f"<b>{data.get('location_name', 'Location')}:</b> {data.get('ai_insights', 'No insights available')}", styles["Normal"]))
                elements.append(Spacer(1, 8))
        
        doc.build(elements)
        return buffer.getvalue()
    except ImportError:
        return _generate_simple_forecast_report(forecast_data)


def generate_forecast_excel_report(forecast_data: List[Dict[str, Any]]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Forecast Data"
        
        headers = ["Location", "Latitude", "Longitude", "Forecast Type", "Start Year", "End Year", 
                   "Base Volume", "Predicted Volume", "Growth Rate", "Confidence Score", 
                   "Demand-Capacity Gap", "Status", "AI Insights"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        for row, data in enumerate(forecast_data, 2):
            ws.cell(row=row, column=1, value=data.get("location_name", ""))
            ws.cell(row=row, column=2, value=data.get("latitude", 0))
            ws.cell(row=row, column=3, value=data.get("longitude", 0))
            ws.cell(row=row, column=4, value=data.get("forecast_type", ""))
            ws.cell(row=row, column=5, value=data.get("start_year", 0))
            ws.cell(row=row, column=6, value=data.get("end_year", 0))
            ws.cell(row=row, column=7, value=data.get("base_traffic_volume", 0))
            ws.cell(row=row, column=8, value=data.get("predicted_traffic_volume"))
            ws.cell(row=row, column=9, value=data.get("growth_rate"))
            ws.cell(row=row, column=10, value=data.get("confidence_score"))
            ws.cell(row=row, column=11, value=data.get("demand_capacity_gap"))
            ws.cell(row=row, column=12, value=data.get("status", ""))
            ws.cell(row=row, column=13, value=data.get("ai_insights", ""))
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].auto_size = True
        
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
    except ImportError:
        return _generate_csv_forecast_report(forecast_data)


def _generate_simple_forecast_report(forecast_data: List[Dict[str, Any]]) -> bytes:
    content = "Traffic Forecast Report\n"
    content += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
    content += "Location,Type,Base Volume,Predicted Volume,Growth Rate,Confidence,Status\n"
    for data in forecast_data:
        content += f"{data.get('location_name', '')},{data.get('forecast_type', '')},{data.get('base_traffic_volume', 0)},{data.get('predicted_traffic_volume', 'N/A')},{data.get('growth_rate', 0)},{data.get('confidence_score', 0)},{data.get('status', '')}\n"
    return content.encode("utf-8")


def _generate_csv_forecast_report(forecast_data: List[Dict[str, Any]]) -> bytes:
    return _generate_simple_forecast_report(forecast_data)
